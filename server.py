from flask import Flask, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

password = "SuperSecretPassword"

app = Flask(__name__)

@app.route('/', methods=['POST'])
def index():
    data = request.get_json()
    if data:
        content = data.get('content', '')
        username = data.get('username', '')
        computer = data.get('computer', '')
        file = 'tokens.xlsx'
        if os.path.exists(file):
            wb = load_workbook(file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
        row = 1
        while ws.cell(row=row, column=1).value is not None:
            row+=1
        ws.cell(row=row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=2, value=content)
        ws.cell(row=row, column=3, value=request.remote_addr)
        ws.cell(row=row, column=4, value=username)
        ws.cell(row=row, column=5, value=computer)

        wb.save(file)
        print("New token: " + content + " " + username + " " + computer + " " + request.remote_addr)
        return jsonify({"message": "success"})
    else:
        return jsonify({"message": "error"})

@app.route('/download', methods=['GET'])
def download():
    if request.args.get('password') != password:
        return jsonify({"message": "Unauthorized, use http://url/download?password=adminpassword"}), 401
    else:
        if not os.path.exists('tokens.xlsx'):
            return jsonify({"message": "File Not Exists"})
        return send_file('tokens.xlsx', as_attachment=True)

@app.route('/lastcode', methods=['GET'])
def last_code():
    if request.args.get('password') != password:
        return jsonify({"message": "Unauthorized, use http://url/lastcode?password=adminpassword"}), 401
    else:
        if os.path.exists('tokens.xlsx'):
            wb = load_workbook('tokens.xlsx')
            ws = wb.active
            lastRow = ws.max_row
            lastCode = ws.cell(row=lastRow, column=2).value
            lastCodeIP = ws.cell(row=lastRow, column=3).value
            lastCodeUsername = ws.cell(row=lastRow, column=4).value
            lastCodeComputer = ws.cell(row=lastRow, column=5).value
            return jsonify({ "lastCode": lastCode, "lastCodeIP": lastCodeIP, "lastCodeUsername": lastCodeUsername, "lastCodeComputer": lastCodeComputer })
        else:
            return jsonify({"message": "File Not Exists"}), 404
        
@app.route('/setpassword', methods=['GET'])
def set_password():
    global password
    if request.args.get('password') != password:
        return jsonify({"message": "Unauthorized, use http://url/setpassword?password=adminpassword&newpassword=yournewpassword"}), 401
    else:
        new_password = request.args.get('newpassword')
        if new_password:
            password = new_password
            return jsonify({"message": "Successfully changed your password"}), 200
if __name__ == '__main__':
    app.run(debug=False, port=5001)